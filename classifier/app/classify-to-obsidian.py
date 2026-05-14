#!/usr/bin/env python3
import argparse
import base64
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional
from zoneinfo import ZoneInfo

import requests

from category_manager import (
    format_examples_for_prompt,
    image_safe_categories,
    load_categories,
    load_relevant_examples,
    normalize_image_classification_result,
    select_candidate_categories,
)

ALASKA_TZ = ZoneInfo("America/Anchorage")

SUPPORTED_EXTENSIONS = {
    ".pdf", ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt",
    ".txt", ".md", ".markdown", ".csv", ".html", ".htm",
    ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"
}

IMAGE_EXTENSIONS = {
    ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".webp"
}

SPREADSHEET_EXTENSIONS = {
    ".xlsx"
}

PLAIN_EXTENSIONS = {
    ".txt", ".md", ".markdown", ".csv"
}

def now_ak() -> str:
    return datetime.now(ALASKA_TZ).isoformat(timespec="seconds")

def elapsed_ms(started_at: float) -> float:
    return round((time.perf_counter() - started_at) * 1000, 3)

def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()

def slugify(value: str, max_len: int = 100) -> str:
    value = re.sub(r"[^A-Za-z0-9._ -]+", "", value)
    value = re.sub(r"\s+", " ", value).strip()
    value = value.replace("/", "-")
    return value[:max_len] or "document"

def obsidian_tag(value: str) -> str:
    value = str(value).lower().strip()
    value = re.sub(r"[^a-z0-9_-]+", "-", value)
    return value.strip("-") or "unknown"

def yaml_list(items: List[Any]) -> str:
    if not items:
        return "[]"
    return "[" + ", ".join(json.dumps(str(x), ensure_ascii=False) for x in items) + "]"

def iter_input_files(path: Path) -> Iterable[Path]:
    if path.is_file():
        if path.suffix.lower() in SUPPORTED_EXTENSIONS:
            yield path
        return

    for p in path.rglob("*"):
        if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS:
            yield p

def ensure_vault(vault: Path) -> None:
    dirs = [
        vault,
        vault / ".obsidian",
        vault / "00 Inbox",
        vault / "01 Classified",
        vault / "02 Needs Review",
        vault / "90 Attachments",
        vault / "_system",
        vault / "_system" / "classifications",
        vault / "_system" / "extracted-markdown",
        vault / "_system" / "templates",
    ]

    for d in dirs:
        d.mkdir(parents=True, exist_ok=True)

    app_json = vault / ".obsidian" / "app.json"
    if not app_json.exists():
        app_json.write_text(json.dumps({
            "alwaysUpdateLinks": True,
            "newFileLocation": "current",
            "attachmentFolderPath": "90 Attachments"
        }, indent=2), encoding="utf-8")

    community_plugins = vault / ".obsidian" / "community-plugins.json"
    if not community_plugins.exists():
        community_plugins.write_text("[]\n", encoding="utf-8")

    home = vault / "Home.md"
    if not home.exists():
        home.write_text("""---
type: vault-home
system: local-document-classifier
---

# Local Document Classifier

Generated Obsidian vault.
""", encoding="utf-8")

def parse_plain_text(path: Path) -> str:
    return path.read_text(encoding="utf-8", errors="replace")

def convert_legacy_office(path: Path, work_dir: Path) -> Optional[Path]:
    ext = path.suffix.lower()
    target_ext = {
        ".doc": "docx",
        ".xls": "xlsx",
        ".ppt": "pptx",
    }.get(ext)

    if not target_ext:
        return None

    work_dir.mkdir(parents=True, exist_ok=True)

    cmd = [
        "soffice",
        "--headless",
        "--convert-to",
        target_ext,
        "--outdir",
        str(work_dir),
        str(path),
    ]

    try:
        subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, timeout=180)
    except Exception:
        return None

    converted = work_dir / f"{path.stem}.{target_ext}"
    return converted if converted.exists() else None

def parse_with_docling(path: Path) -> str:
    from docling.document_converter import DocumentConverter

    converter = DocumentConverter()
    result = converter.convert(str(path))
    return result.document.export_to_markdown()

def parse_spreadsheet_fast(
    path: Path,
    max_sheets: int = 4,
    max_rows: int = 8,
    max_cols: int = 8,
    max_cell_chars: int = 80,
) -> tuple[str, str, Dict[str, Any]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=False, read_only=True)
    sheet_summaries: List[Dict[str, Any]] = []
    non_empty_preview_cells = 0

    for worksheet in workbook.worksheets[:max_sheets]:
        title_lower = worksheet.title.lower()
        is_metadata_sheet = any(token in title_lower for token in ["expectation", "fixture", "metadata"])
        preview_rows: List[List[str]] = []

        if not is_metadata_sheet:
            for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                if row_index > max_rows:
                    break

                values: List[str] = []
                has_value = False

                for cell in row[:max_cols]:
                    value = "" if cell is None else str(cell).strip().replace("\n", " ")
                    value = value[:max_cell_chars]
                    if value:
                        has_value = True
                        non_empty_preview_cells += 1
                    values.append(value)

                if has_value:
                    preview_rows.append(values)

        sheet_summaries.append(
            {
                "title": worksheet.title,
                "preview_rows": preview_rows,
                "is_metadata_sheet": is_metadata_sheet,
            }
        )

    lines = [
        "# Spreadsheet Summary",
        f"Workbook: {path.name}",
        f"Sheet count: {len(workbook.sheetnames)}",
    ]

    for sheet in sheet_summaries:
        lines.append("")
        lines.append(f"## Sheet: {sheet['title']}")

        preview_rows = sheet["preview_rows"]
        if sheet.get("is_metadata_sheet"):
            lines.append("- Metadata-style sheet omitted from classification preview.")
            continue

        if not preview_rows:
            lines.append("- No preview rows with visible values.")
            continue

        for row in preview_rows[:6]:
            cells = [cell for cell in row if cell]
            if cells:
                lines.append(f"- {' | '.join(cells)}")

    markdown = "\n".join(lines).strip()
    metadata = {
        "sheet_count": len(workbook.sheetnames),
        "sheet_names": workbook.sheetnames[:max_sheets],
        "preview_sheet_count": len(sheet_summaries),
        "non_empty_preview_cells": non_empty_preview_cells,
    }
    return markdown, "spreadsheet-openpyxl", metadata

def classify_spreadsheet_fast(
    source_path: Path,
    categories: List[str],
    markdown: Optional[str] = None,
    metadata: Optional[Dict[str, Any]] = None,
) -> tuple[str, Dict[str, Any], Dict[str, Any]]:
    if markdown is None or metadata is None:
        markdown, parser_name, metadata = parse_spreadsheet_fast(source_path)
        metadata["parser"] = parser_name
    else:
        metadata = dict(metadata)
        metadata.setdefault("parser", "spreadsheet-openpyxl")

    text = f"{source_path.name}\n{markdown}".lower()
    available = set(categories)

    def score(*keywords: str) -> int:
        return sum(1 for keyword in keywords if keyword in text)

    domain_scores = {
        "financial": score(
            "budget", "forecast", "variance", "quarter", "fy forecast",
            "actual", "cost", "expense", "revenue", "payment", "department"
        ),
        "tax": score("tax", "irs", "1099", "w-2", "withholding", "deduction"),
        "legal": score("agreement", "contract", "terms", "effective date", "vendor"),
        "medical": score("patient", "medical", "diagnosis", "provider", "prescription"),
        "insurance": score("insurance", "claim", "coverage", "premium", "policy"),
        "technical": score("incident", "network", "server", "uptime", "ticket", "configuration"),
    }

    primary = "spreadsheet" if "spreadsheet" in available else None

    if domain_scores["tax"] >= 2 and "tax" in available:
        primary = "tax"
    elif domain_scores["medical"] >= 2 and "medical" in available:
        primary = "medical"
    elif domain_scores["insurance"] >= 2 and "insurance" in available:
        primary = "insurance"
    elif domain_scores["legal"] >= 2 and "legal" in available:
        primary = "legal"
    elif domain_scores["technical"] >= 3 and "technical" in available:
        primary = "technical"
    elif primary is None and domain_scores["financial"] >= 2 and "financial" in available:
        primary = "financial"

    if primary is None:
        for fallback in ["financial", "work", "report", "unknown", "needs-review"]:
            if fallback in available:
                primary = fallback
                break
        primary = primary or "unknown"

    secondary: List[str] = []
    for label in ["spreadsheet", "financial", "work", "report", "tax", "legal", "medical", "insurance", "technical"]:
        if label != primary and label in available:
            if label == "financial" and domain_scores["financial"] == 0:
                continue
            if label == "technical" and domain_scores["technical"] < 2:
                continue
            if label in {"tax", "legal", "medical", "insurance"} and domain_scores[label] == 0:
                continue
            secondary.append(label)

    secondary = secondary[:4]

    sensitive_flags = [
        label
        for label in ["financial", "tax", "legal", "medical", "insurance"]
        if domain_scores.get(label, 0) > 0
    ] or ["none"]

    confidence = 0.98 if primary == "spreadsheet" else 0.94
    sheet_names = ", ".join(metadata.get("sheet_names", [])[:3]) or "workbook preview"
    reason_parts = [
        "Fast spreadsheet path used to avoid slow OCR/LLM processing.",
        f"Detected workbook structure across {metadata.get('sheet_count', 0)} sheet(s): {sheet_names}.",
    ]
    if domain_scores["financial"] > 0:
        reason_parts.append("Budget and forecast terms indicate financial spreadsheet content.")

    classification = {
        "primary_label": primary,
        "secondary_labels": secondary,
        "confidence": confidence,
        "summary": (
            f"Spreadsheet workbook preview with {metadata.get('sheet_count', 0)} sheet(s), "
            f"including {sheet_names}."
        ),
        "reason": " ".join(reason_parts),
        "sensitive_flags": sensitive_flags,
        "recommended_action": "keep",
        "file_date_guess": "unknown",
        "language": "English",
        "candidate_categories_used": [
            label
            for label in ["spreadsheet", "financial", "work", "report", "tax", "legal", "medical", "insurance", "technical", "unknown", "needs-review"]
            if label in available
        ],
    }
    return markdown, classification, metadata

def parse_document(path: Path, work_dir: Path) -> tuple[str, str]:
    ext = path.suffix.lower()

    if ext in SPREADSHEET_EXTENSIONS:
        markdown, parser_name, _ = parse_spreadsheet_fast(path)
        return markdown, parser_name

    if ext in PLAIN_EXTENSIONS:
        return parse_plain_text(path), "plain-text"

    try:
        return parse_with_docling(path), "docling"
    except Exception as first_error:
        converted = convert_legacy_office(path, work_dir)
        if converted:
            try:
                return parse_with_docling(converted), "docling-converted"
            except Exception:
                pass

        if ext in {".html", ".htm"}:
            return parse_plain_text(path), "html-plain"

        raise RuntimeError(f"Document parsing failed for {path}: {first_error}") from first_error

def extract_json(text: str) -> Dict[str, Any]:
    text = text.strip()

    try:
        return json.loads(text)
    except Exception:
        pass

    match = re.search(r"\{.*\}", text, flags=re.DOTALL)
    if match:
        return json.loads(match.group(0))

    raise ValueError(f"Could not parse JSON from model response: {text[:500]}")

def ollama_chat(
    ollama_url: str,
    model: str,
    messages: List[Dict[str, Any]],
    json_mode: bool = True,
    timeout: int = 600,
) -> str:
    payload: Dict[str, Any] = {
        "model": model,
        "messages": messages,
        "stream": False,
        "options": {
            "temperature": 0
        }
    }

    if json_mode:
        payload["format"] = "json"

    response = requests.post(
        f"{ollama_url.rstrip('/')}/api/chat",
        json=payload,
        timeout=timeout,
    )
    response.raise_for_status()
    data = response.json()
    return data["message"]["content"]

def classify_markdown(
    markdown: str,
    source_path: Path,
    categories: List[str],
    ollama_url: str,
    model: str,
    max_chars: int,
    timing: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    ext = source_path.suffix.lower()
    clipped = markdown[:max_chars]

    candidate_categories = select_candidate_categories(
        all_categories=categories,
        filename=source_path.name,
        extension=ext,
        content=clipped,
        is_image=False,
    )

    examples = load_relevant_examples(
        filename=source_path.name,
        extension=ext,
        content=clipped,
        is_image=False,
    )

    prompt = f"""
You are a private local document classifier.

Classify this file using only the filename, extracted document content, allowed categories, and prior correction examples.
Do not invent names, dates, dollar amounts, account numbers, or legal/medical facts.
If the content is ambiguous, lower confidence and use "unknown" or "needs-review".

Important:
- Choose primary_label only from Allowed categories.
- Use secondary_labels only from Allowed categories.
- A reimbursement packet with attached receipts may be "reimbursement-packet" or "receipt" depending on the dominant content.
- Return strict JSON only.

Allowed categories:
{json.dumps(candidate_categories, indent=2)}

Prior correction examples:
{format_examples_for_prompt(examples)}

Return exactly this JSON shape:
{{
  "primary_label": "one allowed category",
  "secondary_labels": ["zero or more allowed categories"],
  "confidence": 0.0,
  "summary": "short factual summary",
  "reason": "why this label was chosen",
  "sensitive_flags": ["pii", "medical", "financial", "legal", "identity", "none"],
  "recommended_action": "keep, review, archive, reimburse, delete, or unknown",
  "file_date_guess": "YYYY-MM-DD, YYYY-MM, YYYY, or unknown",
  "language": "detected language or unknown"
}}

Filename:
{source_path.name}

Extracted content:
{clipped}
""".strip()

    if timing is not None:
        timing["candidate_category_count"] = len(candidate_categories)
        timing["markdown_chars"] = len(markdown)
        timing["clipped_markdown_chars"] = len(clipped)
        timing["model"] = model

    model_started_at = time.perf_counter()
    content = ollama_chat(
        ollama_url=ollama_url,
        model=model,
        messages=[{"role": "user", "content": prompt}],
        json_mode=True,
    )
    if timing is not None:
        timing["model_ms"] = elapsed_ms(model_started_at)

    result = extract_json(content)
    result["candidate_categories_used"] = candidate_categories
    return result

def classify_image(
    source_path: Path,
    categories: List[str],
    ollama_url: str,
    vision_model: str,
    timing: Optional[Dict[str, Any]] = None,
) -> Dict[str, Any]:
    ext = source_path.suffix.lower()

    candidate_categories = select_candidate_categories(
        all_categories=categories,
        filename=source_path.name,
        extension=ext,
        content="",
        is_image=True,
    )
    candidate_categories = image_safe_categories(candidate_categories, source_path.name)

    examples = load_relevant_examples(
        filename=source_path.name,
        extension=ext,
        content="",
        is_image=True,
    )

    image_b64 = base64.b64encode(source_path.read_bytes()).decode("utf-8")

    prompt = f"""
You are a private local image classifier.

Look at the image and classify it using the allowed categories.

Critical rule:
If it is not a document, do not force it into receipt, legal, medical, insurance, tax, financial, technical, or marketing categories.

For environment/reference images, prefer labels such as reference-image, concept-art, environment-art, game-reference, architecture, industrial, sci-fi, snow-ice, frozen-environment, post-apocalyptic, waystation, facility, building, interior, exterior, machinery, artwork, photo, image-only, or unknown as appropriate.

Describe visible content only.
Do not invent unreadable text.
Return strict JSON only.

Allowed categories:
{json.dumps(candidate_categories, indent=2)}

Prior correction examples:
{format_examples_for_prompt(examples)}

Return exactly this JSON shape:
{{
  "primary_label": "one allowed category",
  "secondary_labels": ["zero or more allowed categories"],
  "confidence": 0.0,
  "summary": "short factual summary of visible content",
  "reason": "why this label was chosen",
  "sensitive_flags": ["pii", "medical", "financial", "legal", "identity", "none"],
  "recommended_action": "keep, review, archive, reimburse, delete, or unknown",
  "file_date_guess": "YYYY-MM-DD, YYYY-MM, YYYY, or unknown",
  "language": "detected language or unknown"
}}

Filename:
{source_path.name}
""".strip()

    if timing is not None:
        timing["candidate_category_count"] = len(candidate_categories)
        timing["image_bytes"] = source_path.stat().st_size
        timing["model"] = vision_model

    model_started_at = time.perf_counter()
    content = ollama_chat(
        ollama_url=ollama_url,
        model=vision_model,
        messages=[{
            "role": "user",
            "content": prompt,
            "images": [image_b64],
        }],
        json_mode=True,
        timeout=900,
    )
    if timing is not None:
        timing["model_ms"] = elapsed_ms(model_started_at)

    result = extract_json(content)
    result["candidate_categories_used"] = candidate_categories
    return normalize_image_classification_result(result)

def write_obsidian_note(
    vault: Path,
    source_path: Path,
    file_hash: str,
    markdown: Optional[str],
    classification: Dict[str, Any],
    attach_originals: bool,
) -> Path:
    # --- image-normalize-before-note BEGIN ---
    try:
        if source_path.suffix.lower() in IMAGE_EXTENSIONS:
            from category_manager import normalize_image_classification_result
            classification = normalize_image_classification_result(classification)
    except Exception:
        pass
    # --- image-normalize-before-note END ---


    primary = str(classification.get("primary_label", "unknown") or "unknown")
    secondary = classification.get("secondary_labels", []) or []
    confidence_raw = classification.get("confidence", 0)
    summary = classification.get("summary", "")
    reason = classification.get("reason", "")
    sensitive_flags = classification.get("sensitive_flags", []) or []
    recommended_action = classification.get("recommended_action", "unknown")
    file_date_guess = classification.get("file_date_guess", "unknown")
    language = classification.get("language", "unknown")

    try:
        confidence = float(confidence_raw)
    except Exception:
        confidence = 0.0

    needs_review = confidence < 0.70 or primary == "unknown"

    if needs_review:
        note_dir = vault / "02 Needs Review"
    else:
        note_dir = vault / "01 Classified" / obsidian_tag(primary)

    note_dir.mkdir(parents=True, exist_ok=True)

    base = f"{slugify(source_path.stem)} - {obsidian_tag(primary)} - {file_hash[:12]}"
    note_path = note_dir / f"{base}.md"

    extracted_link = ""
    if markdown is not None:
        extracted_dir = vault / "_system" / "extracted-markdown" / obsidian_tag(primary)
        extracted_dir.mkdir(parents=True, exist_ok=True)
        extracted_path = extracted_dir / f"{base}.extracted.md"
        extracted_path.write_text(markdown, encoding="utf-8")
        extracted_link = f"[[{extracted_path.relative_to(vault).as_posix()}]]"

    attachment_link = ""
    if attach_originals:
        attachment_dir = vault / "90 Attachments" / obsidian_tag(primary)
        attachment_dir.mkdir(parents=True, exist_ok=True)
        copied = attachment_dir / source_path.name
        if not copied.exists():
            shutil.copy2(source_path, copied)
        attachment_link = f"[[{copied.relative_to(vault).as_posix()}]]"

    tags = [f"classified/{obsidian_tag(primary)}"]

    for item in secondary:
        tags.append(f"classified/{obsidian_tag(item)}")

    if needs_review:
        tags.append("needs-review")

    for flag in sensitive_flags:
        if str(flag).lower() != "none":
            tags.append(f"sensitive/{obsidian_tag(flag)}")

    tags_yaml = "\n".join(f"  - {tag}" for tag in sorted(set(tags)))

    note_body = f"""---
type: classified-document
primary_label: {json.dumps(primary, ensure_ascii=False)}
secondary_labels: {yaml_list(secondary)}
confidence: {confidence}
source_file: {json.dumps(str(source_path), ensure_ascii=False)}
sha256: {json.dumps(file_hash)}
classified_at: {json.dumps(now_ak())}
file_date_guess: {json.dumps(file_date_guess, ensure_ascii=False)}
language: {json.dumps(language, ensure_ascii=False)}
sensitive_flags: {yaml_list(sensitive_flags)}
recommended_action: {json.dumps(recommended_action, ensure_ascii=False)}
attachment: {json.dumps(attachment_link, ensure_ascii=False)}
extracted_markdown: {json.dumps(extracted_link, ensure_ascii=False)}
tags:
{tags_yaml}
---

# {source_path.name}

## Summary

{summary}

## Classification

| Field | Value |
|---|---|
| Primary label | `{primary}` |
| Secondary labels | `{", ".join(map(str, secondary))}` |
| Confidence | `{confidence}` |
| Sensitive flags | `{", ".join(map(str, sensitive_flags))}` |
| Recommended action | `{recommended_action}` |
| File date guess | `{file_date_guess}` |
| Language | `{language}` |
| SHA-256 | `{file_hash}` |

## Reason

{reason}

## Original File

{attachment_link if attachment_link else "Original file was not copied into the vault. Re-run with `--attach-originals` if desired."}

## Extracted Markdown File

{extracted_link if extracted_link else "No extracted Markdown file was written."}

## Extracted Markdown Preview

{markdown[:20000] if markdown else "_No Markdown extraction available for this file._"}
"""

    note_path.write_text(note_body, encoding="utf-8")
    return note_path

def write_index(vault: Path, notes: List[Path]) -> None:
    index = vault / "Classification Index.md"

    lines = [
        "---",
        "type: classification-index",
        "system: local-document-classifier",
        "---",
        "",
        "# Classification Index",
        "",
        f"Last updated: {now_ak()}",
        "",
        "## Recent notes",
        "",
    ]

    for note in notes[-100:]:
        rel = note.relative_to(vault).as_posix()
        lines.append(f"- [[{rel}]]")

    index.write_text("\n".join(lines) + "\n", encoding="utf-8")

def wait_for_ollama(ollama_url: str, timeout_seconds: int = 120) -> None:
    import time

    deadline = time.time() + timeout_seconds
    last_error = None

    while time.time() < deadline:
        try:
            response = requests.get(f"{ollama_url.rstrip('/')}/api/tags", timeout=5)
            if response.ok:
                return
            last_error = f"HTTP {response.status_code}"
        except Exception as e:
            last_error = str(e)

        time.sleep(2)

    raise RuntimeError(f"Ollama did not become ready at {ollama_url}: {last_error}")

def main() -> int:
    parser = argparse.ArgumentParser(description="Dockerized local document classifier that writes to an Obsidian vault.")
    parser.add_argument("input", nargs="?", default="/input", help="Input file or folder inside the container. Default: /input")
    parser.add_argument("--vault", default="/vault", help="Obsidian vault path inside the container. Default: /vault")
    parser.add_argument("--output", default="/output", help="Output folder for manifest/errors. Default: /output")
    parser.add_argument("--work-dir", default="/tmp/work", help="Temporary work directory.")
    parser.add_argument("--ollama-url", default=os.environ.get("OLLAMA_URL", "http://ollama:11434"))
    parser.add_argument("--model", default=os.environ.get("CLASSIFY_MODEL", "qwen2.5:3b"))
    parser.add_argument("--vision-model", default=os.environ.get("VISION_MODEL", "qwen2.5vl:3b"))
    parser.add_argument("--categories", default="__AUTO__")
    parser.add_argument("--max-chars", type=int, default=16000)
    parser.add_argument("--attach-originals", action="store_true")
    parser.add_argument("--no-vision", action="store_true")
    parser.add_argument("--self-test", action="store_true")
    parser.add_argument("--timing-output", default="", help="Optional JSON file path for per-run timing output.")

    args = parser.parse_args()

    input_path = Path(args.input).resolve()
    vault = Path(args.vault).resolve()
    output = Path(args.output).resolve()
    work_dir = Path(args.work_dir).resolve()

    if args.categories == "__AUTO__":
        categories = load_categories()
    else:
        categories = [x.strip() for x in args.categories.split(",") if x.strip()]

    ensure_vault(vault)
    output.mkdir(parents=True, exist_ok=True)
    work_dir.mkdir(parents=True, exist_ok=True)

    wait_for_ollama(args.ollama_url)

    if args.self_test:
        print(json.dumps({
            "ok": True,
            "time_alaska": now_ak(),
            "ollama_url": args.ollama_url,
            "vault": str(vault),
            "output": str(output),
            "model": args.model,
            "vision_model": args.vision_model,
            "category_count": len(categories),
        }, indent=2))
        return 0

    if not input_path.exists():
        print(f"[ERROR] Input path does not exist inside container: {input_path}", file=sys.stderr)
        return 2

    files = list(iter_input_files(input_path))

    if not files:
        print(f"[WARN] No supported files found under: {input_path}")
        return 0

    manifest = output / "manifest.jsonl"
    notes: List[Path] = []
    successes = 0
    failures = 0
    timing_records: List[Dict[str, Any]] = []

    with manifest.open("a", encoding="utf-8") as mf:
        for source_path in files:
            print(f"[INFO] Classifying: {source_path}")
            markdown: Optional[str] = None
            file_hash = ""
            file_started_at = time.perf_counter()
            ext = source_path.suffix.lower()
            timing: Dict[str, Any] = {
                "source_path": str(source_path),
                "filename": source_path.name,
                "extension": ext,
                "mode": "image" if ext in IMAGE_EXTENSIONS and not args.no_vision else "document",
                "attach_originals": args.attach_originals,
                "vision_enabled": not args.no_vision,
            }

            try:
                hash_started_at = time.perf_counter()
                file_hash = sha256_file(source_path)
                timing["sha256_ms"] = elapsed_ms(hash_started_at)

                if ext in IMAGE_EXTENSIONS and not args.no_vision:
                    classification = classify_image(
                        source_path=source_path,
                        categories=categories,
                        ollama_url=args.ollama_url,
                        vision_model=args.vision_model,
                        timing=timing,
                    )
                elif ext in SPREADSHEET_EXTENSIONS:
                    parse_started_at = time.perf_counter()
                    markdown, parser_name, spreadsheet_metadata = parse_spreadsheet_fast(source_path)
                    timing["parse_ms"] = elapsed_ms(parse_started_at)
                    timing["parser"] = parser_name
                    timing["model_ms"] = 0.0
                    timing["classifier"] = "heuristic-spreadsheet-fast-path"
                    timing["candidate_category_count"] = len(
                        [
                            label
                            for label in ["spreadsheet", "financial", "work", "report", "tax", "legal", "medical", "insurance", "technical", "unknown", "needs-review"]
                            if label in categories
                        ]
                    )
                    timing["markdown_chars"] = len(markdown)
                    timing["clipped_markdown_chars"] = len(markdown)
                    timing["sheet_count"] = spreadsheet_metadata.get("sheet_count", 0)
                    markdown, classification, spreadsheet_metadata = classify_spreadsheet_fast(
                        source_path=source_path,
                        categories=categories,
                        markdown=markdown,
                        metadata=spreadsheet_metadata,
                    )
                else:
                    parse_started_at = time.perf_counter()
                    markdown, parser_name = parse_document(source_path, work_dir)
                    timing["parse_ms"] = elapsed_ms(parse_started_at)
                    timing["parser"] = parser_name
                    classification = classify_markdown(
                        markdown=markdown,
                        source_path=source_path,
                        categories=categories,
                        ollama_url=args.ollama_url,
                        model=args.model,
                        max_chars=args.max_chars,
                        timing=timing,
                    )

                note_started_at = time.perf_counter()
                note_path = write_obsidian_note(
                    vault=vault,
                    source_path=source_path,
                    file_hash=file_hash,
                    markdown=markdown,
                    classification=classification,
                    attach_originals=args.attach_originals,
                )
                timing["note_write_ms"] = elapsed_ms(note_started_at)
                timing["primary_label"] = classification.get("primary_label", "unknown")
                timing["secondary_label_count"] = len(classification.get("secondary_labels", []) or [])
                timing["confidence"] = classification.get("confidence")
                timing["ok"] = True
                timing["total_ms"] = elapsed_ms(file_started_at)

                record = {
                    "ok": True,
                    "classified_at": now_ak(),
                    "source_path": str(source_path),
                    "sha256": file_hash,
                    "note_path": str(note_path),
                    "classification": classification,
                    "timing": timing,
                }

                manifest_started_at = time.perf_counter()
                mf.write(json.dumps(record, ensure_ascii=False) + "\n")
                mf.flush()
                timing["manifest_write_ms"] = elapsed_ms(manifest_started_at)
                timing["total_ms"] = elapsed_ms(file_started_at)

                notes.append(note_path)
                timing_records.append(dict(timing))
                successes += 1

                print(f"[OK] {source_path.name} => {classification.get('primary_label', 'unknown')} note={note_path}")

            except Exception as e:
                failures += 1
                timing["ok"] = False
                timing["error"] = str(e)
                timing["total_ms"] = elapsed_ms(file_started_at)

                record = {
                    "ok": False,
                    "classified_at": now_ak(),
                    "source_path": str(source_path),
                    "sha256": file_hash,
                    "error": str(e),
                    "timing": timing,
                }

                mf.write(json.dumps(record, ensure_ascii=False) + "\n")
                mf.flush()
                timing_records.append(dict(timing))

                print(f"[FAIL] {source_path}: {e}", file=sys.stderr)

    write_index(vault, notes)

    if args.timing_output:
        timing_output = Path(args.timing_output)
        timing_output.parent.mkdir(parents=True, exist_ok=True)
        payload: Dict[str, Any]
        if len(timing_records) == 1:
            payload = dict(timing_records[0])
        else:
            payload = {
                "ok": failures == 0,
                "successes": successes,
                "failures": failures,
                "file_count": len(timing_records),
                "files": timing_records,
            }
        timing_output.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"[DONE] successes={successes} failures={failures}")
    print(f"[DONE] vault={vault}")
    print(f"[DONE] manifest={manifest}")

    return 0 if failures == 0 else 1

if __name__ == "__main__":
    raise SystemExit(main())
